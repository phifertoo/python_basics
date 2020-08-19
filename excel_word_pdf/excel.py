import openpyxl
import os

# load_workbook accepts a relative path
workbook = openpyxl.load_workbook('example.xlsx')

# return all of the sheets inside of excel
sheet_list = workbook.get_sheet_names()

# open a sheet
sheet1 = workbook.get_sheet_by_name(sheet_list[0])

# return the cell object within that sheet
cell = sheet1['A1']
a1_cell_value = cell.value
c1_cell_value = sheet1['C1'].value

b2_cell_value = sheet1.cell(row=1, column=3).value

for i in range(1,8):
    print(i, sheet1.cell(row = i, column=2).value)


# create a new workbook
new_wb = openpyxl.Workbook()
sheet1 = new_wb.get_sheet_by_name('Sheet')

# set values in cells
sheet1['A1'].value = 42
sheet1['A2'].value = 'Hello'
new_wb.save('example2.xlsx')

# create new sheets
# we can specify the position
sheet1 = new_wb.create_sheet(index = 0)
sheet1.title = 'my_named_sheet'
new_wb.save('example2.xlsx')

print(new_wb.get_sheet_names())